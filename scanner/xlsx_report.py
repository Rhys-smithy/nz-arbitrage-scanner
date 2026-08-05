"""
Writes an .xlsx version of the scan results with LIVE FORMULAS -- edit the
buyer's premium / platform fee / listing cost assumptions once at the top,
and every row's profit recalculates automatically. This is the per-item
version of the original single-lot landed-cost calculator: same math,
applied to every opportunity from this run in one sheet.

Note: openpyxl writes formulas but not their cached values -- they compute
correctly the moment you actually open this in Excel, Google Sheets, or
Numbers (all recalculate on open), but a tool reading the raw file without
a spreadsheet engine (e.g. a script using data_only=True) would see blank
values until it's been opened at least once in a real app.
"""
import os
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FONT = Font(name=FONT, color="0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
LABEL_FONT = Font(name=FONT, bold=True)
NORMAL_FONT = Font(name=FONT)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("Category", 18),
    ("Source", 12),
    ("Title", 40),
    ("Price (NZD)", 12),
    ("Buy Now (NZD)", 13),
    ("Condition", 12),
    ("Score /10", 9),
    ("Resale Likelihood", 14),
    ("Est. New Price", 13),
    ("Suggested Resale", 14),
    ("Landed Cost", 13),
    ("Net Resale Rev.", 14),
    ("Profit (NZD)", 12),
    ("Profit %", 10),
    ("Notes", 45),
]


def write_xlsx_report(rows: List[Dict]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opportunities"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"] = "NZ Auction Scanner — Opportunities & Profit Estimates"
    ws["A1"].font = TITLE_FONT

    # --- Editable global assumptions (blue text, yellow fill = edit me) ---
    ws["A3"] = "Assumptions (edit these — every row recalculates automatically)"
    ws["A3"].font = LABEL_FONT

    ws["A4"] = "Buyer's premium (%)"
    ws["A4"].font = NORMAL_FONT
    ws["B4"] = 0.15
    ws["B4"].font = INPUT_FONT
    ws["B4"].fill = INPUT_FILL
    ws["B4"].number_format = "0.0%"

    ws["A5"] = "Marketplace/platform fee on resale (%)"
    ws["A5"].font = NORMAL_FONT
    ws["B5"] = 0.10
    ws["B5"].font = INPUT_FONT
    ws["B5"].fill = INPUT_FILL
    ws["B5"].number_format = "0.0%"

    ws["A6"] = "Listing/postage cost per item (NZD)"
    ws["A6"].font = NORMAL_FONT
    ws["B6"] = 10
    ws["B6"].font = INPUT_FONT
    ws["B6"].fill = INPUT_FILL
    ws["B6"].number_format = "#,##0.00"

    ws["A7"] = "Note: figures are Claude's estimates, not real quotes — verify before bidding."
    ws["A7"].font = Font(name=FONT, italic=True, size=9)

    HEADER_ROW = 9
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        cell = ws[f"{col_letter}{HEADER_ROW}"]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    data_start = HEADER_ROW + 1
    for i, row in enumerate(rows):
        r = data_start + i
        price = row.get("price_nzd")
        buy_now = row.get("buy_now_price_nzd")
        resale_price = row.get("suggested_resale_price_nzd")

        ws[f"A{r}"] = row.get("category", "")
        ws[f"B{r}"] = row.get("source", "")
        title_cell = ws[f"C{r}"]
        title_cell.value = row.get("title", "")
        url = row.get("url", "")
        if url:
            title_cell.hyperlink = url
            title_cell.font = Font(name=FONT, color="0563C1", underline="single")
        ws[f"D{r}"] = price if price not in ("", None) else None
        ws[f"E{r}"] = buy_now if buy_now not in ("", None) else None
        ws[f"F{r}"] = row.get("condition", "")
        ws[f"G{r}"] = row.get("score") if row.get("score") not in ("", None) else None
        ws[f"H{r}"] = row.get("resale_likelihood", "")
        ws[f"I{r}"] = row.get("estimated_new_price_nzd") if row.get("estimated_new_price_nzd") not in ("", None) else None
        ws[f"J{r}"] = resale_price if resale_price not in ("", None) else None

        # Landed cost: price + buyer's premium if a bid price exists,
        # otherwise just the buy-now price (no premium on buy-now).
        ws[f"K{r}"] = f'=IF(D{r}<>"",D{r}*(1+$B$4),IF(E{r}<>"",E{r},""))'
        # Net resale revenue: suggested resale minus platform fee minus
        # listing/postage cost.
        ws[f"L{r}"] = f'=IF(J{r}<>"",J{r}*(1-$B$5)-$B$6,"")'
        # Profit and profit %.
        ws[f"M{r}"] = f'=IF(AND(K{r}<>"",L{r}<>""),L{r}-K{r},"")'
        ws[f"N{r}"] = f'=IFERROR(M{r}/K{r},"")'

        ws[f"O{r}"] = row.get("notes", "")

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws[f"{get_column_letter(col_idx)}{r}"]
            cell.border = BORDER
            if col_idx not in (3,):  # title already has its own font
                cell.font = NORMAL_FONT

        for money_col in ("D", "E", "I", "J", "K", "L", "M"):
            ws[f"{money_col}{r}"].number_format = "$#,##0.00;($#,##0.00);-"
        ws[f"N{r}"].number_format = "0.0%"

    path = os.path.join(REPORTS_DIR, f"opportunities_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    os.makedirs(REPORTS_DIR, exist_ok=True)
    wb.save(path)
    return path
